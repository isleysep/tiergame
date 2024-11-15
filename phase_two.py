import cv2 as cv
import numpy as np
import os
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

rankings_df = pd.DataFrame()
directory = os.fsencode("input")
for file in os.listdir(directory):
    print("Reading input from " + os.fsdecode(file))
    # Step 1: Load the train image (the main image where we want to search for templates)
    train_img_path = os.fsdecode(file)
    person_name = os.path.splitext(train_img_path)[0]
    gray = cv.imread("input/"+train_img_path, cv.IMREAD_GRAYSCALE)
    # Step 3: Get the image dimensions
    image_height, image_width = gray.shape

    # Step 4: Known color of the black line (this is an assumption, adjust accordingly)
    # Assuming the black line has a grayscale intensity of 0 (pure black) or a specific RGB value.
    line_color_value = 0  # Change this to the correct value if needed

    # Step 5: Tolerance for detecting "black" (to account for noise)
    tolerance = 40  # Adjust this based on the variation of the black line

    # Step 6: Detect horizontal lines by scanning each row for uniform black pixels
    line_positions = []
    is_line_detected = False  # Flag to check if a line is currently being detected
    for y in range(image_height):
        row = gray[y, :]  # Get all pixels in the row (one row at a time)

        # Check if the row matches the black line color
        if np.all(np.abs(row - line_color_value) <= tolerance):
            if not is_line_detected:
                # Start of a new line block
                line_positions.append(y)
                is_line_detected = True
        else:
            is_line_detected = False  # Reset when row is not part of a line

    # Step 7: Group line positions (in case multiple consecutive rows match the line color)
    filtered_lines = []
    for i in range(1, len(line_positions)):
        if line_positions[i] != line_positions[i-1] + 1:  # New line detected
            filtered_lines.append(line_positions[i-1])
    filtered_lines.append(line_positions[-1])  # Append the last line

    # Step 8: Define bounding boxes between the detected lines
    bounding_boxes = []
    for i in range(len(filtered_lines) - 1):
        y_top = filtered_lines[i]
        y_bottom = filtered_lines[i+1]
        
        # Bounding box coordinates (top left and bottom right corners)
        top_left = (0, y_top)
        bottom_right = (image_width, y_bottom)
        
        bounding_boxes.append((top_left, bottom_right))

    # Step 3: Function to detect features and descriptors using SIFT
    def detect_features(img):
        sift = cv.SIFT_create()
        keypoints, descriptors = sift.detectAndCompute(img, None)
        return keypoints, descriptors

    # Step 4: Function to match descriptors using BFMatcher and the ratio test
    def match_features(des1, des2):
        bf = cv.BFMatcher()
        matches = bf.knnMatch(des1, des2, k=2)
        
        # Apply ratio test
        good_matches = []
        for m, n in matches:
            if m.distance < 0.75 * n.distance:
                good_matches.append(m)
        
        return good_matches

    # Step 5: Function to find the best matching row for a template
    def match_template_to_rows(template_img, bounding_boxes, train_img):
        kp_template, des_template = detect_features(template_img)
        
        best_row = None
        max_good_matches = 0
        img_with_matches = None

        # Step 6: Loop through each bounding box (row)
        for idx, (top_left, bottom_right) in enumerate(bounding_boxes):
            # Crop the bounding box area from the train image
            cropped_region = train_img[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
            
            kp_train, des_train = detect_features(cropped_region)
            
            if des_template is None or des_train is None:
                continue
            
            # Perform feature matching between the template and the cropped region
            good_matches = match_features(des_template, des_train)
            
            # Check if this row has more matches than previous rows
            if len(good_matches) > max_good_matches:
                max_good_matches = len(good_matches)
                best_row = idx

                # Draw matches for the best row
                img_with_matches = cv.drawMatches(template_img, kp_template, cropped_region, kp_train, good_matches, None, flags=cv.DrawMatchesFlags_NOT_DRAW_SINGLE_POINTS)

        return best_row, img_with_matches

    # Step 7: Load all templates from the output folder and match them
    template_folder = 'output/'
    results = {}

    for template_file in os.listdir(template_folder):
        template_path = os.path.join(template_folder, template_file)
        template_img = cv.imread(template_path, cv.IMREAD_GRAYSCALE)

        if template_img is None:
            continue
        
        # Resize template image (if necessary)
        template_img = cv.resize(template_img, (90, 90), interpolation=cv.INTER_LINEAR)
        
        # Find the best matching row for this template
        best_row, img_with_matches = match_template_to_rows(template_img, bounding_boxes, gray)
        
        if best_row is not None:
            results[template_file] = best_row
            
            # # Show the matched result
            # plt.imshow(img_with_matches)
            # plt.title(f'Matched {template_file} in Row {best_row}')
            # plt.show()

    # Step 8: Output the results
    for template, row in results.items():
        template = os.path.splitext(template)[0]
        if person_name not in rankings_df.columns:
            rankings_df[person_name] = None
        if template not in rankings_df.index:
            rankings_df.loc[template] = [None] * len(rankings_df.columns)
        rankings_df.at[template, person_name] = 6 - row

# Add "Average Rating" and "Standard Deviation" columns, rounded to 2 decimal places
rankings_df["Average Rating"] = rankings_df.mean(axis=1).round(2)
rankings_df["Standard Deviation"] = rankings_df.iloc[:, :-1].std(axis=1).round(2)  # Exclude "Average Rating" column

# Reorder columns to place "Average Rating" as the first column
columns = ["Average Rating"] + [col for col in rankings_df.columns if col not in ["Average Rating", "Standard Deviation"]] + ["Standard Deviation"]
rankings_df = rankings_df[columns]

# Add a row for each person's average rating, rounded to 2 decimal places
average_row = pd.DataFrame(rankings_df.iloc[:, 1:-1].mean(axis=0).round(2)).T  # Calculate the average rating for each person
average_row.index = ["Average Rating"]
rankings_df = pd.concat([rankings_df, average_row])

# Save the DataFrame to an Excel file
rankings_df.index.name = "Song Name"
excel_filename = "song_rankings.xlsx"
rankings_df.to_excel(excel_filename, sheet_name="Rankings")

# Load the workbook and apply conditional formatting for the background colors
wb = load_workbook(excel_filename)
ws = wb["Rankings"]

# Set column width to approximately 1 inch (Excel unit = 10.71)
for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 10.71

# Define colors for average and standard deviation gradients
def get_color_gradient(value, min_val, max_val, start_color, end_color):
    """Get the hex color based on the value within a range for the gradient."""
    def interpolate_color(val, min_val, max_val, start_color, end_color):
        fraction = (val - min_val) / (max_val - min_val)
        return int(start_color + (end_color - start_color) * fraction)
    
    start_rgb = [int(start_color[i:i+2], 16) for i in (0, 2, 4)]
    end_rgb = [int(end_color[i:i+2], 16) for i in (0, 2, 4)]
    
    interpolated_rgb = [
        interpolate_color(value, min_val, max_val, start_rgb[i], end_rgb[i])
        for i in range(3)
    ]
    
    return f"{interpolated_rgb[0]:02X}{interpolated_rgb[1]:02X}{interpolated_rgb[2]:02X}"

min_val, max_val = 0, 6
start_color, end_color = "FFFFFF", "FF5050"

# Apply background color based on the value for each cell, average rating, and standard deviation
for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.value = round(cell.value, 2)  # Ensure value is rounded to 2 decimal places
            color = get_color_gradient(cell.value, min_val, max_val, start_color, end_color)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Save the updated workbook
wb.save(excel_filename)